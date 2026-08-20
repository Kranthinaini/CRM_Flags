import io
from datetime import date

import pandas as pd
import streamlit as st


# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="CRM Productivity Abnormality Checker",
    page_icon="🔎",
    layout="wide",
)

st.title("🔎 CRM Productivity – Abnormality Checker")

st.caption(
    "Upload the CRM Productivity Dump and identify abnormal records."
)


# ============================================================
# CONFIGURATION
# ============================================================

FAKE_GPS_LAT = 51.673858
FAKE_GPS_LON = 7.815982


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def find_column(df, candidates):
    """
    Return the first matching column,
    ignoring case and extra spaces.
    """

    normalized = {
        str(c).strip().lower(): c
        for c in df.columns
    }

    for candidate in candidates:

        key = candidate.strip().lower()

        if key in normalized:
            return normalized[key]

    return None


def clean_mobile(value):
    """
    Convert Excel/string mobile values to digits.
    """

    if pd.isna(value):
        return ""

    text = str(value).strip()

    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]

    return "".join(
        ch for ch in text
        if ch.isdigit()
    )


def parse_datetime(df):
    """
    Create LeadDateTime from Lead Date + Lead Time.
    """

    date_col = find_column(
        df,
        ["Lead Date"]
    )

    time_col = find_column(
        df,
        ["Lead Time"]
    )

    if not date_col:

        return pd.Series(
            pd.NaT,
            index=df.index
        )

    lead_date = pd.to_datetime(
        df[date_col],
        errors="coerce"
    )

    if time_col:

        time_text = (
            df[time_col]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        combined = (
            lead_date.dt.strftime("%Y-%m-%d")
            + " "
            + time_text
        )

        parsed = pd.to_datetime(
            combined,
            errors="coerce"
        )

        parsed = parsed.fillna(
            lead_date
        )

        return parsed

    return lead_date


# ============================================================
# ANALYZE CRM
# ============================================================

def analyze_crm(
    df,
    quick_minutes=3
):

    df = df.copy()

    df.columns = [
        str(c).strip()
        for c in df.columns
    ]

    # --------------------------------------------------------
    # REQUIRED COLUMNS
    # --------------------------------------------------------

    required = [
        "Lead Date",
        "Lead Time",
        "EmployeeName",
        "Mobile Number",
        "Latitude",
        "Longitude",
    ]

    missing = [
        c for c in required
        if c not in df.columns
    ]

    if missing:

        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing)
        )

    # --------------------------------------------------------
    # LEAD DATE
    # --------------------------------------------------------

    df["Lead Date"] = pd.to_datetime(
        df["Lead Date"],
        errors="coerce"
    ).dt.date

    # --------------------------------------------------------
    # LEAD DATE + TIME
    # --------------------------------------------------------

    lead_dt = parse_datetime(df)

    # --------------------------------------------------------
    # MOBILE CHECKS
    # --------------------------------------------------------

    mobiles = (
        df["Mobile Number"]
        .apply(clean_mobile)
    )

    placeholder_mobile = mobiles.eq(
        "0000000000"
    )

    invalid_mobile = (
        mobiles.ne("")
        & ~placeholder_mobile
        & mobiles.str.len().ne(10)
    )

    # Blank mobile is not separately flagged
    # because original logic did not include it
    blank_mobile = mobiles.eq("")

    # --------------------------------------------------------
    # GPS CHECKS
    #
    # NOTE: GPS values are compared at FULL PRECISION
    # (no rounding). Two shops located physically close
    # to each other will have slightly different raw
    # lat/lon readings and must NOT be treated as the
    # same location just because they are nearby.
    # Only an EXACT match (identical lat AND identical
    # lon) counts as a duplicate.
    # --------------------------------------------------------

    lat = pd.to_numeric(
        df["Latitude"],
        errors="coerce"
    )

    lon = pd.to_numeric(
        df["Longitude"],
        errors="coerce"
    )

    # Fake GPS (exact match against known spoofed default)
    fake_gps = (
        lat.eq(FAKE_GPS_LAT)
        & lon.eq(FAKE_GPS_LON)
    )

    # GPS key (exact raw values, not rounded)
    gps_key = (
        lat.astype("string")
        + "|"
        + lon.astype("string")
    )

    # Duplicate GPS
    duplicate_gps = (
        gps_key.map(
            gps_key.value_counts()
        ).gt(1)
        & lat.notna()
        & lon.notna()
        & ~fake_gps
    )

    # --------------------------------------------------------
    # QUICK VISIT CHECK
    # --------------------------------------------------------
    #
    # IMPORTANT:
    #
    # Quick visits are checked separately
    # for:
    #
    # EmployeeName + Lead Date
    #
    # Therefore:
    #
    # 06-Aug visit will NOT be compared
    # with 07-Aug visit.
    #
    # --------------------------------------------------------

    temp = pd.DataFrame(
        {
            "EmployeeName": (
                df["EmployeeName"]
                .fillna("")
                .astype(str)
                .str.strip()
            ),

            "LeadDate": df["Lead Date"],

            "LeadDateTime": lead_dt,

            "OriginalIndex": df.index,
        }
    )

    temp = temp.sort_values(
        [
            "EmployeeName",
            "LeadDate",
            "LeadDateTime",
            "OriginalIndex",
        ]
    )

    gaps = (
        temp
        .groupby(
            [
                "EmployeeName",
                "LeadDate"
            ],
            dropna=False
        )["LeadDateTime"]
        .diff()
        .dt.total_seconds()
        .div(60)
    )

    quick_indices = temp.loc[
        gaps.lt(quick_minutes)
        & gaps.ge(0),
        "OriginalIndex",
    ]

    quick_visit = pd.Series(
        False,
        index=df.index
    )

    quick_visit.loc[
        quick_indices
    ] = True

    # --------------------------------------------------------
    # BUILD FLAGS
    # --------------------------------------------------------

    summaries = []
    totals = []

    for idx in df.index:

        flags = []

        # ----------------------------------------------------
        # MOBILE
        # ----------------------------------------------------

        if placeholder_mobile.loc[idx]:

            flags.append(
                "Placeholder mobile (0000000000)"
            )

        elif invalid_mobile.loc[idx]:

            flags.append(
                "Invalid mobile format"
            )

        # ----------------------------------------------------
        # GPS
        # ----------------------------------------------------

        if fake_gps.loc[idx]:

            flags.append(
                "Fake GPS (Germany default)"
            )

        elif duplicate_gps.loc[idx]:

            flags.append(
                "Duplicate GPS w/ other visit"
            )

        # ----------------------------------------------------
        # QUICK VISIT
        # ----------------------------------------------------

        if quick_visit.loc[idx]:

            flags.append(
                f"Visit <{quick_minutes} min "
                "after previous (same emp)"
            )

        # ----------------------------------------------------
        # INVALID GPS
        # ----------------------------------------------------

        if (
            pd.isna(lat.loc[idx])
            or pd.isna(lon.loc[idx])
        ):

            flags.append(
                "Invalid / missing GPS"
            )

        # ----------------------------------------------------
        # TOTAL FLAGS
        # ----------------------------------------------------

        totals.append(
            len(flags)
        )

        summaries.append(
            "; ".join(flags)
        )

    # --------------------------------------------------------
    # RESULT
    # --------------------------------------------------------

    result = df.copy()

    result["Lead Date"] = pd.to_datetime(
        result["Lead Date"],
        errors="coerce"
    ).dt.date

    result["Total_Flags"] = totals

    result["Flag_Summary"] = summaries

    # --------------------------------------------------------
    # FH / SH CLASSIFICATION
    # --------------------------------------------------------
    # FH = before 2:00 PM
    # SH = 2:00 PM onwards
    # This uses Lead Date + Lead Time already parsed above.
    result["_LeadDateTime"] = lead_dt
    result["Session"] = result["_LeadDateTime"].apply(
        lambda x: (
            "FH" if pd.notna(x) and x.hour < 14
            else "SH" if pd.notna(x)
            else ""
        )
    )

    # --------------------------------------------------------
    # FH / SH EMPLOYEE ERROR SUMMARY
    # --------------------------------------------------------
    # Fake GPS = Technical Error
    # Every other abnormality = CM Error.
    # If a visit has both Fake GPS and another flag, it contributes
    # to both Technical Error and CM Error.
    # Only employees with at least one error (Total Flags > 0) are
    # included in this summary.
    def build_session_error_summary(session):
        session_df = result[result["Session"] == session].copy()

        if session_df.empty:
            return pd.DataFrame(
                columns=[
                    "EmployeeName",
                    "Total Visits",
                    "Total Flags",
                    "Technical Error",
                    "CM Error",
                    "Remarks",
                ]
            )

        rows = []

        for employee, group in session_df.groupby(
            "EmployeeName",
            dropna=False
        ):
            total_visits = len(group)
            total_flags = int(group["Total_Flags"].sum())

            # Skip employees with no errors at all in this session.
            if total_flags == 0:
                continue

            technical_error = 0
            cm_error = 0
            remarks = []

            for _, visit in group.iterrows():
                flag_text = str(visit.get("Flag_Summary", "") or "")
                if not flag_text:
                    continue

                flags = [
                    f.strip()
                    for f in flag_text.split(";")
                    if f.strip()
                ]

                has_fake_gps = any(
                    "Fake GPS" in f
                    for f in flags
                )

                other_flags = [
                    f for f in flags
                    if "Fake GPS" not in f
                ]

                if has_fake_gps:
                    technical_error += 1

                if other_flags:
                    cm_error += len(other_flags)

                for flag in flags:
                    if flag not in remarks:
                        remarks.append(flag)

            rows.append(
                {
                    "EmployeeName": employee,
                    "Total Visits": total_visits,
                    "Total Flags": total_flags,
                    "Technical Error": technical_error,
                    "CM Error": cm_error,
                    "Remarks": " / ".join(remarks),
                }
            )

        if not rows:
            return pd.DataFrame(
                columns=[
                    "EmployeeName",
                    "Total Visits",
                    "Total Flags",
                    "Technical Error",
                    "CM Error",
                    "Remarks",
                ]
            )

        return (
            pd.DataFrame(rows)
            .sort_values(
                ["Total Flags", "EmployeeName"],
                ascending=[False, True],
                na_position="last",
            )
            .reset_index(drop=True)
        )

    fh_error_summary = build_session_error_summary("FH")
    sh_error_summary = build_session_error_summary("SH")

    # --------------------------------------------------------
    # ABNORMALITY SUMMARY
    # --------------------------------------------------------

    summary_rows = [

        (
            "Placeholder mobile (0000000000)",
            int(
                placeholder_mobile.sum()
            )
        ),

        (
            "Invalid mobile format",
            int(
                invalid_mobile.sum()
            )
        ),

        (
            "Fake GPS (Germany default)",
            int(
                fake_gps.sum()
            )
        ),

        (
            "Duplicate GPS w/ other visit",
            int(
                duplicate_gps.sum()
            )
        ),

        (
            f"Visit <{quick_minutes} min "
            "after previous (same emp/date)",
            int(
                quick_visit.sum()
            )
        ),

        (
            "Invalid / missing GPS",
            int(
                (
                    lat.isna()
                    | lon.isna()
                ).sum()
            )
        ),
    ]

    summary = pd.DataFrame(
        summary_rows,
        columns=[
            "Abnormality",
            "Count"
        ]
    )

    # --------------------------------------------------------
    # EMPLOYEE SUMMARY
    # --------------------------------------------------------

    emp_col = "EmployeeName"

    employee_summary = (
        result
        .groupby(
            emp_col,
            dropna=False
        )
        .agg(
            Total_Visits=(
                emp_col,
                "size"
            ),

            Flagged_Visits=(
                "Total_Flags",
                lambda x: int(
                    (x > 0).sum()
                )
            ),

            Total_Flags=(
                "Total_Flags",
                "sum"
            ),
        )
        .reset_index()
        .sort_values(
            [
                "Total_Flags",
                "Flagged_Visits"
            ],
            ascending=False
        )
    )

    return (
        result,
        summary,
        employee_summary,
        fh_error_summary,
        sh_error_summary
    )


# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel(
    result,
    summary,
    employee_summary,
    fh_error_summary,
    sh_error_summary,
    from_date,
    to_date
):

    output = io.BytesIO()

    # --------------------------------------------------------
    # DATE RANGE TEXT
    # --------------------------------------------------------

    if from_date == to_date:

        date_text = from_date.strftime(
            "%Y-%m-%d"
        )

    else:

        date_text = (
            f"{from_date.strftime('%Y-%m-%d')}"
            f" to "
            f"{to_date.strftime('%Y-%m-%d')}"
        )

    # --------------------------------------------------------
    # OVERVIEW
    # --------------------------------------------------------

    overview = pd.DataFrame(
        {
            "Metric": [
                "Report Date Range",
                "Total CRM Records",
                "Flagged Records",
                "Total Flags",
            ],

            "Value": [
                date_text,

                len(result),

                int(
                    (
                        result["Total_Flags"]
                        > 0
                    ).sum()
                ),

                int(
                    result["Total_Flags"].sum()
                ),
            ],
        }
    )

    # --------------------------------------------------------
    # WRITE EXCEL
    # --------------------------------------------------------

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        overview.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

        summary.to_excel(
            writer,
            sheet_name="Abnormality Summary",
            index=False
        )

        employee_summary.to_excel(
            writer,
            sheet_name="Employee Summary",
            index=False
        )

        # ----------------------------------------------------
        # FH & SH ERROR SUMMARY - SINGLE SHEET, TWO TABLES
        # ----------------------------------------------------
        # FH table starts at row 1 (header at row 1).
        # SH table starts a couple of rows below the end of
        # the FH table, each with its own title + header row.

        fh_title_row = 0  # 0-indexed row for "FH Error Summary" title
        fh_header_row = 1  # header row for FH table
        fh_data_start = 2  # first data row for FH table

        # startrow is 0-indexed for to_excel; title goes one row above header
        fh_error_summary.to_excel(
            writer,
            sheet_name="FH & SH Error Summary",
            index=False,
            startrow=fh_header_row,
        )

        sh_title_row = (
            fh_header_row
            + 1
            + len(fh_error_summary)
            + 2
        )
        sh_header_row = sh_title_row + 1

        sh_error_summary.to_excel(
            writer,
            sheet_name="FH & SH Error Summary",
            index=False,
            startrow=sh_header_row,
        )

        result_for_excel = result.drop(
            columns=["_LeadDateTime", "Session"],
            errors="ignore"
        )

        result_for_excel.to_excel(
            writer,
            sheet_name="Flagged CRM Data",
            index=False
        )

        # ----------------------------------------------------
        # FORMATTING
        # ----------------------------------------------------

        ws = writer.book[
            "Flagged CRM Data"
        ]

        from openpyxl.styles import (
            PatternFill,
            Font,
            Alignment
        )

        # ----------------------------------------------------
        # LEAD DATE FORMAT
        # ----------------------------------------------------

        header_map_initial = {
            str(cell.value).strip():
                cell.column
            for cell in ws[1]
        }

        lead_date_col = (
            header_map_initial.get(
                "Lead Date"
            )
        )

        if lead_date_col:

            for row_num in range(
                2,
                ws.max_row + 1
            ):

                cell = ws.cell(
                    row_num,
                    lead_date_col
                )

                if isinstance(
                    cell.value,
                    date
                ):

                    cell.number_format = (
                        "dd-mm-yyyy"
                    )

        # ----------------------------------------------------
        # COLORS
        # ----------------------------------------------------

        header_fill = PatternFill(
            "solid",
            fgColor="1F4E78"
        )

        abnormal_fill = PatternFill(
            "solid",
            fgColor="FCE4D6"
        )

        # ----------------------------------------------------
        # HEADER
        # ----------------------------------------------------

        for cell in ws[1]:

            cell.fill = header_fill

            cell.font = Font(
                color="FFFFFF",
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # ----------------------------------------------------
        # FLAG COLUMNS
        # ----------------------------------------------------

        header_map = {
            str(cell.value).strip():
                cell.column
            for cell in ws[1]
        }

        total_flags_col = (
            header_map.get(
                "Total_Flags"
            )
        )

        flag_summary_col = (
            header_map.get(
                "Flag_Summary"
            )
        )

        # ----------------------------------------------------
        # HIGHLIGHT FLAGGED ROWS
        # ----------------------------------------------------

        if total_flags_col:

            for row_num in range(
                2,
                ws.max_row + 1
            ):

                value = ws.cell(
                    row_num,
                    total_flags_col
                ).value

                try:

                    has_flag = (
                        float(value or 0)
                        > 0
                    )

                except (
                    TypeError,
                    ValueError
                ):

                    has_flag = False

                if has_flag:

                    for col_num in range(
                        1,
                        ws.max_column + 1
                    ):

                        ws.cell(
                            row_num,
                            col_num
                        ).fill = abnormal_fill

        # ----------------------------------------------------
        # TOTAL FLAGS ALIGNMENT
        # ----------------------------------------------------

        if total_flags_col:

            for row_num in range(
                2,
                ws.max_row + 1
            ):

                ws.cell(
                    row_num,
                    total_flags_col
                ).alignment = Alignment(
                    horizontal="center",
                    vertical="top"
                )

        # ----------------------------------------------------
        # FLAG SUMMARY WRAP
        # ----------------------------------------------------

        if flag_summary_col:

            for row_num in range(
                2,
                ws.max_row + 1
            ):

                ws.cell(
                    row_num,
                    flag_summary_col
                ).alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        # ----------------------------------------------------
        # FREEZE + FILTER
        # ----------------------------------------------------

        ws.freeze_panes = "A2"

        ws.auto_filter.ref = (
            ws.dimensions
        )

        # ----------------------------------------------------
        # FH & SH ERROR SUMMARY FORMATTING (single sheet,
        # two stacked tables with their own titles)
        # ----------------------------------------------------

        combo_ws = writer.book["FH & SH Error Summary"]

        widths = {
            "EmployeeName": 25,
            "Total Visits": 14,
            "Total Flags": 13,
            "Technical Error": 17,
            "CM Error": 12,
            "Remarks": 55,
        }

        def format_error_table(title_row_0idx, header_row_0idx, n_rows, title_text):
            title_row = title_row_0idx + 1  # 1-indexed for openpyxl
            header_row = header_row_0idx + 1

            # Title cell
            title_cell = combo_ws.cell(title_row, 1, title_text)
            title_cell.font = Font(bold=True, size=13, color="1F4E78")

            # Header row styling
            header_cells = combo_ws[header_row]
            header_map_local = {}

            for cell in header_cells:

                if cell.value is None:
                    continue

                header_map_local[str(cell.value).strip()] = cell.column

                cell.fill = header_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

            data_start = header_row + 1
            data_end = header_row + n_rows

            for column_name in [
                "Total Visits",
                "Total Flags",
                "Technical Error",
                "CM Error",
            ]:
                col_num = header_map_local.get(column_name)

                if col_num:
                    for row_num in range(data_start, data_end + 1):
                        combo_ws.cell(row_num, col_num).alignment = Alignment(
                            horizontal="center",
                            vertical="center"
                        )

            remarks_col = header_map_local.get("Remarks")

            if remarks_col:
                for row_num in range(data_start, data_end + 1):
                    combo_ws.cell(row_num, remarks_col).alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

            for header, width in widths.items():
                col_num = header_map_local.get(header)
                if col_num:
                    combo_ws.column_dimensions[
                        combo_ws.cell(header_row, col_num).column_letter
                    ].width = width

        format_error_table(
            fh_title_row,
            fh_header_row,
            len(fh_error_summary),
            "FH Error Summary (before 2:00 PM)",
        )

        format_error_table(
            sh_title_row,
            sh_header_row,
            len(sh_error_summary),
            "SH Error Summary (2:00 PM onwards)",
        )

    # Remove helper columns from the downloadable detailed data.
    result = result.drop(
        columns=["_LeadDateTime", "Session"],
        errors="ignore"
    )

    output.seek(0)

    return output.getvalue()


# ============================================================
# SIDEBAR
# ============================================================

st.sidebar.header("⚙️ Settings")

quick_minutes = st.sidebar.number_input(
    "Quick visit threshold (minutes)",
    min_value=1,
    max_value=30,
    value=3,
    step=1,
)

st.sidebar.info(
    "Upload the CRM dump and select a "
    "custom Lead Date range. The app checks "
    "mobile numbers, GPS, duplicate locations, "
    "quick visits and missing GPS."
)


# ============================================================
# FILE UPLOAD
# ============================================================

uploaded = st.file_uploader(
    "📂 Upload CRM Productivity Dump",
    type=[
        "xlsx",
        "xls",
        "csv"
    ],
)


if uploaded is None:

    st.info(
        "Upload the CRM Excel file to start."
    )

    st.stop()


# ============================================================
# READ FILE
# ============================================================

try:

    if uploaded.name.lower().endswith(
        ".csv"
    ):

        df = pd.read_csv(
            uploaded
        )

        selected_sheet = "CSV"

    else:

        xls = pd.ExcelFile(
            uploaded
        )

        if len(xls.sheet_names) == 1:

            selected_sheet = (
                xls.sheet_names[0]
            )

        else:

            selected_sheet = (
                st.selectbox(
                    "Select CRM sheet",
                    xls.sheet_names
                )
            )

        df = pd.read_excel(
            uploaded,
            sheet_name=selected_sheet
        )

except Exception as exc:

    st.error(
        f"Could not read the uploaded file: {exc}"
    )

    st.stop()


# ============================================================
# CLEAN COLUMN NAMES
# ============================================================

df.columns = [
    str(c).strip()
    for c in df.columns
]


# ============================================================
# FIND LEAD DATE COLUMN
# ============================================================

lead_date_col = find_column(
    df,
    ["Lead Date"]
)

if not lead_date_col:

    st.error(
        "The uploaded file does not contain "
        "a 'Lead Date' column."
    )

    st.stop()


# ============================================================
# CONVERT LEAD DATE
# ============================================================

df[lead_date_col] = pd.to_datetime(
    df[lead_date_col],
    errors="coerce"
).dt.date


# ============================================================
# AVAILABLE DATES
# ============================================================

unique_dates = sorted(
    df[lead_date_col]
    .dropna()
    .unique()
)


if not unique_dates:

    st.error(
        "No valid Lead Date values were found."
    )

    st.stop()


min_available_date = min(
    unique_dates
)

max_available_date = max(
    unique_dates
)


# ============================================================
# CUSTOM DATE RANGE
# ============================================================

st.subheader(
    "📅 Select CRM Date Range"
)

col1, col2 = st.columns(2)


with col1:

    from_date = st.date_input(
        "From Date",
        value=min_available_date,
        min_value=min_available_date,
        max_value=max_available_date,
        format="DD-MM-YYYY",
    )


with col2:

    to_date = st.date_input(
        "To Date",
        value=max_available_date,
        min_value=min_available_date,
        max_value=max_available_date,
        format="DD-MM-YYYY",
    )


# ============================================================
# VALIDATE DATE RANGE
# ============================================================

if from_date > to_date:

    st.error(
        "From Date cannot be greater than To Date."
    )

    st.stop()


# ============================================================
# FILTER DATE RANGE
# ============================================================

df_daily = df[
    df[lead_date_col].between(
        from_date,
        to_date
    )
].copy()


# ============================================================
# SELECTED DATES
# ============================================================

selected_dates = sorted(
    df_daily[lead_date_col]
    .dropna()
    .unique()
)


# ============================================================
# DATE RANGE INFORMATION
# ============================================================

if from_date == to_date:

    selected_date_text = (
        from_date.strftime(
            "%d-%m-%Y"
        )
    )

else:

    selected_date_text = (
        f"{from_date.strftime('%d-%m-%Y')}"
        f" to "
        f"{to_date.strftime('%d-%m-%Y')}"
    )


st.success(
    f"Selected Date Range: "
    f"{selected_date_text} "
    f"| Records: {len(df_daily):,}"
)


# ============================================================
# ANALYZE CRM
# ============================================================

try:

    (
        result,
        summary,
        employee_summary,
        fh_error_summary,
        sh_error_summary,
    ) = analyze_crm(
        df_daily,
        quick_minutes=quick_minutes
    )

except Exception as exc:

    st.error(
        str(exc)
    )

    st.stop()


# ============================================================
# DASHBOARD (KEY METRICS ONLY)
# ============================================================
# Detailed breakdowns (Abnormality Summary, Flagged CRM Data,
# Employee Summary, FH/SH Error Summary) are available in the
# downloadable Excel report below instead of being repeated
# on-screen.

total_records = len(result)

flagged_records = int(
    (
        result["Total_Flags"]
        > 0
    ).sum()
)

total_flags = int(
    result["Total_Flags"].sum()
)

clean_records = (
    total_records
    - flagged_records
)


c1, c2, c3, c4 = st.columns(4)


c1.metric(
    "Total Records",
    f"{total_records:,}"
)

c2.metric(
    "Flagged Records",
    f"{flagged_records:,}"
)

c3.metric(
    "Total Flags",
    f"{total_flags:,}"
)

c4.metric(
    "Clean Records",
    f"{clean_records:,}"
)


# ============================================================
# CREATE EXCEL
# ============================================================

excel_bytes = create_excel(
    result,
    summary,
    employee_summary,
    fh_error_summary,
    sh_error_summary,
    from_date,
    to_date
)


# ============================================================
# DOWNLOAD FILE NAME
# ============================================================

download_name = (
    f"CRM_Abnormality_Report_"
    f"{from_date.strftime('%Y%m%d')}_to_"
    f"{to_date.strftime('%Y%m%d')}.xlsx"
)


# ============================================================
# DOWNLOAD
# ============================================================

st.download_button(
    "⬇️ Download Abnormality Report",
    data=excel_bytes,
    file_name=download_name,
    mime=(
        "application/"
        "vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet"
    ),
    type="primary",
)


# ============================================================
# FOOTER
# ============================================================

st.caption(
    "Full breakdown (Abnormality Summary, Flagged CRM Data, "
    "Employee Summary, FH & SH Error Summary) is in the "
    "downloadable Excel report. Quick visits are "
    "checked separately for each employee "
    "and each Lead Date. FH is before 2:00 PM and "
    "SH is 2:00 PM onwards. Fake GPS is classified "
    "as Technical Error; all other abnormalities are "
    "classified as CM Error. FH & SH Error Summary tables "
    "list only employees who have at least one error."
)